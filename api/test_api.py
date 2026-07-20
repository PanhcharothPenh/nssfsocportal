import urllib.request
import urllib.parse
import json
import os
import openpyxl
import traceback

WORKSPACE = r"c:\Users\Miller\Documents\SOC-Work-WebAPP"

def run_tests():
    print("Starting API integration tests...")
    
    # 1. Test Dashboard
    try:
        req = urllib.request.urlopen("http://127.0.0.1:8000/api/dashboard")
        data = json.loads(req.read().decode('utf-8'))
        print("OK: Dashboard API check successful.")
    except Exception as e:
        print(f"FAIL: Dashboard API check failed: {e}")
        return
        
    # Get a valid branch ID dynamically
    try:
        req_list = urllib.request.urlopen("http://127.0.0.1:8000/api/branches")
        branches_list = json.loads(req_list.read().decode('utf-8'))
        first_branch_id = branches_list[0]['id']
        print(f"Using dynamic Branch ID = {first_branch_id} for testing.")
    except Exception as e:
        print(f"FAIL: Fetch branches list failed: {e}")
        return

    # 2. Test Branch Detail
    try:
        req = urllib.request.urlopen(f"http://127.0.0.1:8000/api/branches/{first_branch_id}")
        data = json.loads(req.read().decode('utf-8'))
        print("OK: Branch detail API check successful.")
    except Exception as e:
        print(f"FAIL: Branch detail API check failed: {e}")
        return
        
    # 3. Test IP Update Sync (Post to first_branch_id, IP 192.168.141.12)
    test_ip = "192.168.141.12"
    payload = {
        "user_name": "Test User",
        "position": "Tech Staff",
        "mac_address": "AA:BB:CC:DD:EE:FF",
        "device_type": "Laptop",
        "status": "Using",
        "internet_permission": "Allow",
        "other": "Integration Test Entry"
    }
    
    try:
        data_bytes = json.dumps(payload).encode('utf-8')
        req_url = f"http://127.0.0.1:8000/api/branches/{first_branch_id}/ips?ip={test_ip}"
        req = urllib.request.Request(
            req_url, 
            data=data_bytes, 
            headers={'Content-Type': 'application/json'}
        )
        with urllib.request.urlopen(req) as response:
            res_data = json.loads(response.read().decode('utf-8'))
            print("OK: Update IP API request successful.")
    except Exception as e:
        print(f"FAIL: Update IP API request failed: {e}")
        return
        
    # 4. Verify in Excel directly using openpyxl
    try:
        file_path = os.path.join(WORKSPACE, "IP Address Branch.xlsx")
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_name = "១.ខណ្ឌដង្កោ"
        sheet = wb[sheet_name]
        
        found = False
        for r in range(2, sheet.max_row + 1):
            ip_val = str(sheet.cell(row=r, column=2).value or '').strip()
            if ip_val == test_ip:
                name_val = str(sheet.cell(row=r, column=3).value or '').strip()
                mac_val = str(sheet.cell(row=r, column=5).value or '').strip()
                print("OK: Excel direct verify check successful.")
                assert name_val == "Test User"
                assert mac_val == "AA:BB:CC:DD:EE:FF"
                found = True
                break
        if not found:
            print(f"FAIL: Excel direct verify check failed: IP {test_ip} not found in sheet")
            return
    except Exception as e:
        print(f"FAIL: Excel direct verify check failed: {e}")
        return
        
    # 5. Clean up entry
    clean_payload = {
        "user_name": None,
        "position": None,
        "mac_address": None,
        "device_type": None,
        "status": "Available",
        "internet_permission": None,
        "other": None
    }
    try:
        data_bytes = json.dumps(clean_payload).encode('utf-8')
        req_url = f"http://127.0.0.1:8000/api/branches/{first_branch_id}/ips?ip={test_ip}"
        req = urllib.request.Request(
            req_url, 
            data=data_bytes, 
            headers={'Content-Type': 'application/json'}
        )
        with urllib.request.urlopen(req) as response:
            res_data = json.loads(response.read().decode('utf-8'))
            print("OK: Cleanup IP API request successful.")
    except Exception as e:
        print(f"FAIL: Cleanup IP API request failed: {e}")
        return

    # 6. Test Hospital Reopen & Reference Doc sync
    try:
        req_list = urllib.request.urlopen("http://127.0.0.1:8000/api/hospital_vpns")
        hospitals = json.loads(req_list.read().decode('utf-8'))
        target_hos = None
        for h in hospitals:
            if "សហមេត្រី" in h['name'] and h['vpn_type'] == 'Close':
                target_hos = h
                break
                
        if not target_hos:
            print("FAIL: Target hospital for reopen test not found in DB")
            return
            
        print(f"Found target hospital: ID={target_hos['id']}")
        
        # Post reopen request
        reopen_payload = {
            "name": target_hos["name"],
            "address": target_hos["address"],
            "isp": target_hos["isp"],
            "public_ip": target_hos["public_ip"],
            "status": "Completed",
            "other": target_hos["other"],
            "reopen_requested": 1,
            "reference_doc": "លិខិតយោងលេខ ៩៩៩ ប.ស.ស"
        }
        
        data_bytes = json.dumps(reopen_payload).encode('utf-8')
        req_url = f"http://127.0.0.1:8000/api/hospital_vpns/{target_hos['id']}"
        req = urllib.request.Request(
            req_url,
            data=data_bytes,
            headers={'Content-Type': 'application/json'}
        )
        with urllib.request.urlopen(req) as response:
            res_data = json.loads(response.read().decode('utf-8'))
            print("OK: Reopen request API successful.")
            assert res_data["status"] == "success"
            
        # Verify in Excel directly
        file_path = os.path.join(WORKSPACE, "តារាងមន្ទីរពេទ្យឯកជន-PRIVATE-HOSPITAL-VPN-2025.xlsx")
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_name = "VPN-HOS-Close"
        sheet = wb[sheet_name]
        
        found = False
        for r in range(2, sheet.max_row + 1):
            name_val = str(sheet.cell(row=r, column=2).value or '').strip()
            if "សហមេត្រី" in name_val:
                other_val = str(sheet.cell(row=r, column=6).value or '').strip()
                print("OK: Excel direct verify closed hospital successful.")
                print(f"DEBUG other_val matches assertions: {'បើកវិញបណ្តោះអាសន្ន' in other_val and 'លិខិតយោងលេខ ៩៩៩' in other_val}")
                assert "បើកវិញបណ្តោះអាសន្ន" in other_val
                assert "លិខិតយោងលេខ ៩៩៩" in other_val
                found = True
                break
                
        if not found:
            print("FAIL: Reopen Excel sync check failed: target row not found")
            return
            
        # Reset back to original values
        reset_payload = {
            "name": target_hos["name"],
            "address": target_hos["address"],
            "isp": target_hos["isp"],
            "public_ip": target_hos["public_ip"],
            "status": "Completed",
            "other": "បើកវិញបណ្តោះអាសន្ន 22-6-2026",
            "reopen_requested": 1,
            "reference_doc": "បើកវិញបណ្តោះអាសន្ន 22-6-2026"
        }
        data_bytes = json.dumps(reset_payload).encode('utf-8')
        req = urllib.request.Request(
            req_url,
            data=data_bytes,
            headers={'Content-Type': 'application/json'}
        )
        with urllib.request.urlopen(req) as response:
            res_data = json.loads(response.read().decode('utf-8'))
            print("OK: Reset reopen request successful.")
            assert res_data["status"] == "success"
            
    except Exception as e:
        print("FAIL: Hospital Reopen test failed:")
        traceback.print_exc()
        return
        
    print("\n===========================================")
    print("ALL API AND EXCEL INTEGRATION TESTS PASSED!")
    print("===========================================")

if __name__ == "__main__":
    run_tests()
