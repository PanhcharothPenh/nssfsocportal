import os
import openpyxl
import sqlite3
import glob

WORKSPACE = r"c:\Users\Miller\Documents\SOC-Work-WebAPP"

def get_db_connection():
    DB_PATH = os.path.join(WORKSPACE, "soc_network.db")
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def sync_branch_ip_to_excel(branch_id, ip_address, updates):
    """
    Updates the IP entry in IP Address Branch.xlsx
    updates: dict of field -> new_value
    """
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Get branch sheet name
    cursor.execute("SELECT name_kh, sheet_name FROM branches WHERE id = ?", (branch_id,))
    branch = cursor.fetchone()
    if not branch or not branch['sheet_name']:
        conn.close()
        return False, "Branch sheet not found"
        
    sheet_name = branch['sheet_name']
    file_path = os.path.join(WORKSPACE, "IP Address Branch.xlsx")
    
    if not os.path.exists(file_path):
        conn.close()
        return False, "Branch Excel file not found"
        
    try:
        wb = openpyxl.load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            conn.close()
            return False, f"Sheet '{sheet_name}' not found in workbook"
            
        sheet = wb[sheet_name]
        
        # Find IP address column and row
        ip_col_idx = None
        header_row_idx = None
        
        # Search first 10 rows for headers
        for r_idx in range(1, 11):
            row_vals = [str(sheet.cell(row=r_idx, column=c).value or '').lower().replace(' ', '') for c in range(1, sheet.max_column + 1)]
            if any('ipaddress' in val or 'ip' in val for val in row_vals):
                header_row_idx = r_idx
                # find column index (1-based)
                for c_idx in range(1, sheet.max_column + 1):
                    val = str(sheet.cell(row=r_idx, column=c_idx).value or '').lower().replace(' ', '')
                    if 'ipaddress' in val or 'ip' in val:
                        ip_col_idx = c_idx
                        break
                break
                
        if not ip_col_idx or not header_row_idx:
            # Fallback search of all cells in first 10 rows
            for r_idx in range(1, 11):
                for c_idx in range(1, sheet.max_column + 1):
                    val = str(sheet.cell(row=r_idx, column=c_idx).value or '')
                    if '.' in val and (val.startswith('192.') or val.startswith('172.') or val.startswith('10.')):
                        # Seems like an IP row data. Header is likely above it.
                        ip_col_idx = c_idx
                        header_row_idx = max(1, r_idx - 1)
                        break
                if ip_col_idx:
                    break
                    
        if not ip_col_idx:
            conn.close()
            return False, "Could not locate IP column in sheet"
            
        # Get headers to map fields to columns
        headers = {}
        for c_idx in range(1, sheet.max_column + 1):
            header_val = str(sheet.cell(row=header_row_idx, column=c_idx).value or '').lower().replace(' ', '').replace('-', '')
            if header_val:
                headers[header_val] = c_idx
                
        # Find the row containing our target IP
        target_row_idx = None
        for r_idx in range(header_row_idx + 1, sheet.max_row + 1):
            cell_val = str(sheet.cell(row=r_idx, column=ip_col_idx).value or '').strip()
            if cell_val == ip_address.strip():
                target_row_idx = r_idx
                break
                
        if not target_row_idx:
            conn.close()
            return False, f"IP Address {ip_address} not found in sheet"
            
        # Update sheet cells
        # Field mapping for Branch: user_name, position, mac_address, device_type, status, internet_permission, other
        col_mappings = {
            'user_name': ['username', 'name', 'ឈ្មោះ'],
            'position': ['position', 'pos', 'ឋានៈ', 'តួនាទី'],
            'mac_address': ['mac', 'macaddress', 'physical'],
            'device_type': ['device', 'devicetype', 'type', 'ប្រភេទ'],
            'status': ['status', 'ស្ថានភាព'],
            'internet_permission': ['perm', 'internet', 'internetpermission'],
            'other': ['other', 'ផ្សេង', 'ផ្សេងៗ']
        }
        
        for field, new_val in updates.items():
            if field not in col_mappings:
                continue
            # Find column index for this field
            target_col_idx = None
            possible_headers = col_mappings[field]
            for h_key, h_col in headers.items():
                if any(ph in h_key for ph in possible_headers):
                    target_col_idx = h_col
                    break
            
            if target_col_idx:
                sheet.cell(row=target_row_idx, column=target_col_idx).value = new_val
                
        wb.save(file_path)
        conn.close()
        
        # Google Sheets Sync
        try:
            from google_sheets import sync_ip_to_google_sheet
            sync_ip_to_google_sheet('branch', sheet_name, ip_address, updates)
        except Exception as gs_err:
            print(f"Google Sheets Branch Sync Error: {gs_err}")
            
        return True, "Synced successfully"
        
    except Exception as e:
        if conn:
            conn.close()
        return False, str(e)

def sync_hq_ip_to_excel(dept_id, ip_address, updates):
    """
    Updates the IP entry in NSSF HQ Users IP Address 2024.xlsx
    updates: dict of field -> new_value
    """
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Get department sheet name
    cursor.execute("SELECT name_en, sheet_name FROM hq_departments WHERE id = ?", (dept_id,))
    dept = cursor.fetchone()
    if not dept or not dept['sheet_name']:
        conn.close()
        return False, "Department sheet not found"
        
    sheet_name = dept['sheet_name']
    file_path = os.path.join(WORKSPACE, "NSSF HQ Users IP Address 2024.xlsx")
    
    if not os.path.exists(file_path):
        conn.close()
        return False, "HQ Excel file not found"
        
    try:
        wb = openpyxl.load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            conn.close()
            return False, f"Sheet '{sheet_name}' not found in workbook"
            
        sheet = wb[sheet_name]
        
        # Find IP Address column
        ip_col_idx = None
        header_row_idx = None
        
        # Search first 10 rows for headers
        for r_idx in range(1, 11):
            row_vals = [str(sheet.cell(row=r_idx, column=c).value or '').lower().replace(' ', '').replace('_', '') for c in range(1, sheet.max_column + 1)]
            if any('ipaddress' in val or 'ip' in val for val in row_vals):
                header_row_idx = r_idx
                # find column index (1-based)
                for c_idx in range(1, sheet.max_column + 1):
                    val = str(sheet.cell(row=r_idx, column=c_idx).value or '').lower().replace(' ', '').replace('_', '')
                    if 'ipaddress' in val or 'ip' in val:
                        ip_col_idx = c_idx
                        break
                break
                
        if not ip_col_idx:
            conn.close()
            return False, "Could not locate IP column in sheet"
            
        # Get headers
        headers = {}
        for c_idx in range(1, sheet.max_column + 1):
            header_val = str(sheet.cell(row=header_row_idx, column=c_idx).value or '').lower().replace(' ', '').replace('-', '').replace('_', '')
            if header_val:
                headers[header_val] = c_idx
                
        # Check sub-headers if there's Latin/Khmer usernames
        sub_headers = {}
        if header_row_idx + 1 <= sheet.max_row:
            for c_idx in range(1, sheet.max_column + 1):
                sub_val = str(sheet.cell(row=header_row_idx + 1, column=c_idx).value or '').lower().strip()
                if sub_val in ['latin', 'khmer']:
                    sub_headers[c_idx] = sub_val
                    
        # Find target row
        target_row_idx = None
        for r_idx in range(header_row_idx + 1, sheet.max_row + 1):
            cell_val = str(sheet.cell(row=r_idx, column=ip_col_idx).value or '').strip()
            if cell_val == ip_address.strip():
                target_row_idx = r_idx
                break
                
        if not target_row_idx:
            conn.close()
            return False, f"IP Address {ip_address} not found in sheet"
            
        # Update sheet cells
        # Field mapping for HQ: user_name_kh, user_name_en, position, old_ip, subnet_mask, gateway, status, internet_permission, group_system, verify_update, other
        col_mappings = {
            'user_name_kh': ['username', 'name', 'khmer', 'ឈ្មោះ'],
            'user_name_en': ['username', 'name', 'latin', 'ឈ្មោះ'],
            'position': ['position', 'pos', 'ឋានៈ', 'តួនាទី'],
            'old_ip': ['oldip'],
            'subnet_mask': ['subnetmask', 'mask'],
            'gateway': ['gateway', 'gw'],
            'status': ['status', 'ស្ថានភាព'],
            'internet_permission': ['permission', 'internet', 'internetpermission'],
            'group_system': ['group', 'groupsystem', 'systemgroup'],
            'verify_update': ['verify', 'verifyupdate', 'lastday'],
            'other': ['other', 'ផ្សេង', 'ផ្សេងៗ']
        }
        
        for field, new_val in updates.items():
            if field not in col_mappings:
                continue
                
            target_col_idx = None
            possible_headers = col_mappings[field]
            
            # Special logic for latin/khmer splits
            if field == 'user_name_kh':
                # find column where sub_header is khmer, or header has khmer
                for c_idx, sub_val in sub_headers.items():
                    if sub_val == 'khmer':
                        target_col_idx = c_idx
                        break
                if not target_col_idx:
                    for h_key, h_col in headers.items():
                        if 'khmer' in h_key:
                            target_col_idx = h_col
                            break
            elif field == 'user_name_en':
                for c_idx, sub_val in sub_headers.items():
                    if sub_val == 'latin':
                        target_col_idx = c_idx
                        break
                if not target_col_idx:
                    for h_key, h_col in headers.items():
                        if 'latin' in h_key:
                            target_col_idx = h_col
                            break
                            
            # General search if not resolved by sub-header
            if not target_col_idx:
                for h_key, h_col in headers.items():
                    if any(ph in h_key for ph in possible_headers):
                        target_col_idx = h_col
                        break
                        
            if target_col_idx:
                sheet.cell(row=target_row_idx, column=target_col_idx).value = new_val
                
        wb.save(file_path)
        conn.close()
        
        # Google Sheets Sync
        try:
            from google_sheets import sync_ip_to_google_sheet
            sync_ip_to_google_sheet('hq', sheet_name, ip_address, updates)
        except Exception as gs_err:
            print(f"Google Sheets HQ Sync Error: {gs_err}")
            
        return True, "Synced successfully"
        
    except Exception as e:
        if conn:
            conn.close()
        return False, str(e)

def sync_vpn_user_to_excel(vpn_id, updates):
    """
    Updates the VPN Remote Access user
    """
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT no, username, name FROM vpn_remote_users WHERE id = ?", (vpn_id,))
    user = cursor.fetchone()
    if not user:
        conn.close()
        return False, "VPN User not found in DB"
        
    db_no = user['no']
    db_username = user['username']
    db_name = user['name']
    conn.close()
    
    files = glob.glob(os.path.join(WORKSPACE, "*VPN Remote Access*.xlsx"))
    if not files:
        return False, "VPN Excel file not found"
    file_path = files[0]
    
    try:
        wb = openpyxl.load_workbook(file_path)
        synced = False
        
        # Search both sheets
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Find username or name column
            user_col_idx = None
            name_col_idx = None
            header_row_idx = None
            
            for r_idx in range(1, 10):
                row_vals = [str(sheet.cell(row=r_idx, column=c).value or '').lower().replace(' ', '') for c in range(1, sheet.max_column + 1)]
                if 'username' in row_vals or 'គោត្តនាម' in row_vals:
                    header_row_idx = r_idx
                    for c_idx in range(1, sheet.max_column + 1):
                        val = str(sheet.cell(row=r_idx, column=c_idx).value or '').lower().replace(' ', '')
                        if 'username' in val:
                            user_col_idx = c_idx
                        elif 'គោត្តនាម' in val or 'name' in val:
                            name_col_idx = c_idx
                    break
                    
            if not header_row_idx:
                continue
                
            # Headers map
            headers = {}
            for c_idx in range(1, sheet.max_column + 1):
                header_val = str(sheet.cell(row=header_row_idx, column=c_idx).value or '').lower().replace(' ', '').replace('-', '')
                if header_val:
                    headers[header_val] = c_idx
                    
            # Find row
            target_row_idx = None
            for r_idx in range(header_row_idx + 1, sheet.max_row + 1):
                u_val = str(sheet.cell(row=r_idx, column=user_col_idx).value or '').strip() if user_col_idx else ''
                n_val = str(sheet.cell(row=r_idx, column=name_col_idx).value or '').strip() if name_col_idx else ''
                
                # Check match
                match = False
                if db_username and u_val == db_username:
                    match = True
                elif db_name and n_val == db_name:
                    match = True
                elif db_no and str(sheet.cell(row=r_idx, column=1).value or '').strip() == str(db_no):
                    match = True
                    
                if match:
                    target_row_idx = r_idx
                    break
                    
            if target_row_idx:
                col_mappings = {
                    'name': ['name', 'គោត្តនាម'],
                    'position': ['position', 'ឋានៈ'],
                    'username': ['username'],
                    'password': ['password'],
                    'department': ['department'],
                    'company': ['company', 'ប.ស.ស'],
                    'status': ['status', 'ស្ថានភាព'],
                    'purpose': ['purpose', 'គោលបំណង'],
                    'vpn_type': ['vpn'],
                    'other': ['other', 'ផ្សេង']
                }
                
                for field, new_val in updates.items():
                    if field not in col_mappings:
                        continue
                    target_col_idx = None
                    possible_headers = col_mappings[field]
                    for h_key, h_col in headers.items():
                        if any(ph in h_key for ph in possible_headers):
                            target_col_idx = h_col
                            break
                    if target_col_idx:
                        sheet.cell(row=target_row_idx, column=target_col_idx).value = new_val
                
                synced = True
                
        if synced:
            wb.save(file_path)
            return True, "Synced successfully"
        else:
            return False, "User not found in Excel sheets"
            
    except Exception as e:
        return False, str(e)

def sync_hospital_vpn_to_excel(vpn_id, updates):
    """
    Updates the Hospital S2S VPN details in PRIVATE-HOSPITAL-VPN-2025.xlsx
    """
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT no, name, public_ip, vpn_type FROM hospital_vpns WHERE id = ?", (vpn_id,))
    vpn = cursor.fetchone()
    if not vpn:
        conn.close()
        return False, "VPN Configuration not found in DB"
        
    db_no = vpn['no']
    db_name = vpn['name']
    db_ip = vpn['public_ip']
    db_type = vpn['vpn_type']
    conn.close()
    
    files = glob.glob(os.path.join(WORKSPACE, "*PRIVATE-HOSPITAL-VPN*.xlsx"))
    if not files:
        return False, "Hospital VPN Excel file not found"
    file_path = files[0]
    
    # Map database vpn_type to Sheet name
    sheet_name = None
    if db_type == 'S2S':
        sheet_name = 'Hospital-VPN-S2S'
    elif db_type == 'Close':
        sheet_name = 'VPN-HOS-Close'
    elif db_type == 'Bank':
        sheet_name = 'Bank-VPN'
        
    if not sheet_name:
        return False, "Invalid VPN Type"
        
    try:
        wb = openpyxl.load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            return False, f"Sheet '{sheet_name}' not found in workbook"
            
        sheet = wb[sheet_name]
        
        # Locate header row
        header_row_idx = None
        for r_idx in range(1, 10):
            row_vals = [str(sheet.cell(row=r_idx, column=c).value or '').lower().replace(' ', '') for c in range(1, sheet.max_column + 1)]
            if any('hospitalname' in val or 'ippublic' in val or 'status' in val or 'publicip' in val for val in row_vals):
                header_row_idx = r_idx
                break
                
        if not header_row_idx:
            header_row_idx = 1
            
        # Headers map
        headers = {}
        for c_idx in range(1, sheet.max_column + 1):
            header_val = str(sheet.cell(row=header_row_idx, column=c_idx).value or '').lower().replace(' ', '').replace('-', '')
            if header_val:
                headers[header_val] = c_idx
                
        # Find row by matching name or no or public IP
        target_row_idx = None
        for r_idx in range(header_row_idx + 1, sheet.max_row + 1):
            row_no = str(sheet.cell(row=r_idx, column=1).value or '').strip()
            # check hospital name column (usually column 2)
            row_name = str(sheet.cell(row=r_idx, column=2).value or '').strip()
            
            match = False
            if db_no and row_no == str(db_no):
                match = True
            elif db_name and row_name.lower() == db_name.lower():
                match = True
                
            if match:
                target_row_idx = r_idx
                break
                
        if not target_row_idx:
            return False, "VPN row not found in Excel sheet"
            
        # Field mapping depends on vpn_type
        if db_type == 'S2S':
            # Fields: name, address, isp, public_ip, subnet, gateway, lan_ip, lan_subnet, lan_gateway, ikey, tunnel, status, contact, year, device, other
            # S2S sheet has duplicate "Subnet", "Gateway" names
            # We track them by index order from headers
            subnet_cols = [c_idx for h_key, c_idx in headers.items() if 'subnet' in h_key]
            gw_cols = [c_idx for h_key, c_idx in headers.items() if 'gateway' in h_key]
            
            # Query full DB record to construct Other text correctly (combining reference_doc and other)
            conn_db = get_db_connection()
            db_cursor = conn_db.cursor()
            db_cursor.execute("SELECT reference_doc, other FROM hospital_vpns WHERE id = ?", (vpn_id,))
            db_row = db_cursor.fetchone()
            conn_db.close()
            
            db_ref = db_row['reference_doc'] if db_row else ''
            db_other = db_row['other'] if db_row else ''
            
            # Format other cell value
            final_other_val = ""
            if db_ref and db_other:
                final_other_val = f"{db_ref} | {db_other}"
            elif db_ref:
                final_other_val = db_ref
            else:
                final_other_val = db_other

            for field, new_val in updates.items():
                if field == 'subnet' and len(subnet_cols) > 0:
                    sheet.cell(row=target_row_idx, column=subnet_cols[0]).value = new_val
                elif field == 'lan_subnet' and len(subnet_cols) > 1:
                    sheet.cell(row=target_row_idx, column=subnet_cols[1]).value = new_val
                elif field == 'gateway' and len(gw_cols) > 0:
                    sheet.cell(row=target_row_idx, column=gw_cols[0]).value = new_val
                elif field == 'lan_gateway' and len(gw_cols) > 1:
                    sheet.cell(row=target_row_idx, column=gw_cols[1]).value = new_val
                elif field in ['other', 'reference_doc']:
                    target_col_idx = None
                    for h_key, h_col in headers.items():
                        if 'other' in h_key:
                            target_col_idx = h_col
                            break
                    if target_col_idx:
                        sheet.cell(row=target_row_idx, column=target_col_idx).value = final_other_val
                else:
                    col_mappings = {
                        'name': ['hospitalname'],
                        'address': ['address'],
                        'isp': ['isp'],
                        'public_ip': ['publicip'],
                        'lan_ip': ['lanip'],
                        'ikey': ['ikey'],
                        'tunnel': ['tunnel'],
                        'status': ['status'],
                        'contact': ['contact'],
                        'year': ['year'],
                        'device': ['device']
                    }
                    if field in col_mappings:
                        possible_headers = col_mappings[field]
                        target_col_idx = None
                        for h_key, h_col in headers.items():
                            if any(ph in h_key for ph in possible_headers):
                                target_col_idx = h_col
                                break
                        if target_col_idx:
                            sheet.cell(row=target_row_idx, column=target_col_idx).value = new_val
                            
        elif db_type == 'Close':
            # Columns: No, Hospital Name, Address, Year, Status, Other
            # Query full DB record to construct Other text correctly
            conn = get_db_connection()
            db_cursor = conn.cursor()
            db_cursor.execute("SELECT reopen_requested, reference_doc, other FROM hospital_vpns WHERE id = ?", (vpn_id,))
            db_row = db_cursor.fetchone()
            conn.close()
            
            db_reopen = db_row['reopen_requested'] if db_row else 0
            db_ref = db_row['reference_doc'] if db_row else ''
            db_other = db_row['other'] if db_row else ''
            
            # Format other cell value
            final_other_val = db_ref or db_other
            if db_reopen:
                if final_other_val and "បើកវិញ" not in final_other_val and "ស្នើសុំ" not in final_other_val:
                    final_other_val = f"បើកវិញបណ្តោះអាសន្ន {final_other_val}"
                elif not final_other_val:
                    final_other_val = "បើកវិញបណ្តោះអាសន្ន"
            
            for field, new_val in updates.items():
                col_mappings = {
                    'name': ['hospitalname'],
                    'address': ['address'],
                    'year': ['year'],
                    'status': ['status'],
                    'other': ['other']
                }
                if field in col_mappings:
                    possible_headers = col_mappings[field]
                    target_col_idx = None
                    for h_key, h_col in headers.items():
                        if any(ph in h_key for ph in possible_headers):
                            target_col_idx = h_col
                            break
                    if target_col_idx:
                        if field == 'other':
                            sheet.cell(row=target_row_idx, column=target_col_idx).value = final_other_val
                        else:
                            sheet.cell(row=target_row_idx, column=target_col_idx).value = new_val
                            
            # Write other if reopen_requested or reference_doc was updated but 'other' was not explicitly passed
            if 'other' not in updates and ('reopen_requested' in updates or 'reference_doc' in updates):
                other_col_idx = None
                for h_key, h_col in headers.items():
                    if 'other' in h_key or 'ផ្សេង' in h_key:
                        other_col_idx = h_col
                        break
                if other_col_idx:
                    sheet.cell(row=target_row_idx, column=other_col_idx).value = final_other_val
                        
        elif db_type == 'Bank':
            # Columns: #, Name, IP Public, IP Address PRO, NSSF-IP-PRO, Bank IP STG/SIT, Bank IP UAT, IKE Version, Pre-Share Key, Device, Date
            for field, new_val in updates.items():
                col_mappings = {
                    'name': ['name'],
                    'public_ip': ['ippublic'],
                    'lan_ip': ['ipaddresspro', 'pro'],
                    'ikey': ['presharekey', 'preshare'],
                    'device': ['device'],
                    'status': ['ikeversion', 'ike'],
                }
                # Other is complex, we skip syncing 'other' directly or overwrite if formatted
                if field in col_mappings:
                    possible_headers = col_mappings[field]
                    target_col_idx = None
                    for h_key, h_col in headers.items():
                        if any(ph in h_key for ph in possible_headers):
                            target_col_idx = h_col
                            break
                    if target_col_idx:
                        sheet.cell(row=target_row_idx, column=target_col_idx).value = new_val
                        
        wb.save(file_path)
        
        # Google Sheets Sync
        try:
            from google_sheets import sync_hospital_vpn_to_google_sheet
            sync_hospital_vpn_to_google_sheet(sheet_name, db_type, db_no, db_name, updates)
        except Exception as gs_err:
            print(f"Google Sheets Hospital Sync Error: {gs_err}")
            
        return True, "Synced successfully"
        
    except Exception as e:
        return False, str(e)
